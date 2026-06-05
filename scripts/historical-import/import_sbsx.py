#!/usr/bin/env python3
"""
One-shot importer for the Showtime Bowl Series X historical Excel export.

Reads the StatisticDownload_*.xlsx file, pivots the long stat rows into the
shape that POST /admin/matches/{id}/import expects, looks up each match in
the live DB via the admin API, and posts the data.

This whole `scripts/historical-import/` folder is throwaway — delete it once
the migration is done.

Usage:
    export SHOWTIME_TOKEN="<paste your admin bearer token from devtools>"
    export SHOWTIME_API_URL="http://localhost:8089/api/v1"  # optional, this is the default
    python3 import_sbsx.py <path-to-xlsx>
    python3 import_sbsx.py <path-to-xlsx> --dry-run     # parse & match without POSTing
    python3 import_sbsx.py <path-to-xlsx> --only "Sun 24 Sep 2023"   # only matches on a date

    # Workbook with one sheet per season (e.g. Showtime Stats.xlsx):
    python3 import_sbsx.py <path-to-xlsx> \
        --sheet "Season XI" --competition "Showtime Bowl Series XI" --dry-run

Token: log into the admin UI, open devtools → Application → Local Storage,
copy `showtime_access_token`.
"""

import argparse
import json
import os
import re
import sys
from collections import defaultdict
from datetime import datetime

import openpyxl
import urllib.request
import urllib.error

# ---- Config ------------------------------------------------------------------

COMPETITION_NAME = "Showtime Bowl Series X"

# Excel "Stat Desc" -> API field. Apps isn't a stat, it's used to mark sheet presence.
STAT_MAP = {
    "Reception": "receptions",
    "Flag Pulls": "flag_pulls",
    "Drops": "drops",
    "Receiving TDs": "receiving_tds",
    "Pass Deflections": "pass_deflections",
    "Passing Attempts": "passing_attempts",
    "Interceptions": "interceptions",
    "Completed Passes": "completed_passes",
    "Passing TDs": "passing_tds",
    "QB Sacks": "qb_sacks",
    "Extra Points TDs": "extra_points_tds",
    "Interception Thrown": "interceptions_thrown",
    "Defensive TDs": "defensive_tds",
    "Safety": "safety",
}

# Excel team name (lowercased, after stripping the (CODE) suffix) -> DB team name.
# Only listed for teams whose Excel name differs from the DB name; identical ones
# don't need entries. The DB names below come from the live roster JSON.
TEAM_NAME_OVERRIDES: dict[str, str] = {
    "abia-warriors":     "Abia Warriors",
    "delta-panthers":    "Delta Panthers",
    "lagos mavericks":   "Mavericks",
    "outlaws athletics": "Outlaws",
    "spartans fc":       "Spartans",
    "titans athletics":  "Titans",
    "cross river ikan":  "Cross Rivers Ikan Sports",
}

GENDER_SUFFIX_RE = re.compile(r"\s*\((M|F)\)\s*$", re.IGNORECASE)
TEAM_CODE_RE = re.compile(r"\s*\([A-Z]+\)\s*$")


# ---- HTTP helpers ------------------------------------------------------------

def api_request(method: str, path: str, token: str, base_url: str, body=None):
    url = f"{base_url}{path}"
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(url, data=data, method=method)
    req.add_header("Authorization", f"Bearer {token}")
    req.add_header("Content-Type", "application/json")
    try:
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        body_text = e.read().decode(errors="replace")
        raise SystemExit(f"API {method} {path} -> {e.code}: {body_text}")


# ---- Excel parsing -----------------------------------------------------------

def strip_gender(name: str) -> str:
    return GENDER_SUFFIX_RE.sub("", name or "").strip()

def strip_team_code(team: str) -> str:
    return TEAM_CODE_RE.sub("", team or "").strip()

def parse_fixture_date(s: str) -> datetime:
    # "Sun 24 Sep 2023 11:00 AM"
    return datetime.strptime(s, "%a %d %b %Y %I:%M %p")

def read_competition_rows(xlsx_path: str, competition_name: str, sheet: str | None = None) -> list[dict]:
    """
    Pull rows whose Fixture Group equals competition_name.

    If `sheet` is given, only that sheet is read.
    Otherwise: prefer the legacy `StatisticDownload_20260506_1319` sheet for
    backward compatibility, then auto-detect by scanning every sheet's
    Fixture Group column. This lets a single workbook hold many seasons.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if sheet:
        if sheet not in wb.sheetnames:
            raise SystemExit(f"Sheet {sheet!r} not found. Available: {wb.sheetnames}")
        sheets_to_check = [sheet]
    elif "StatisticDownload_20260506_1319" in wb.sheetnames:
        sheets_to_check = ["StatisticDownload_20260506_1319"]
    else:
        sheets_to_check = list(wb.sheetnames)

    rows = []
    for sheet_name in sheets_to_check:
        ws = wb[sheet_name]
        header = [c.value for c in ws[1]]
        if "Fixture Group" not in header:
            continue
        for r in ws.iter_rows(min_row=2, values_only=True):
            rec = dict(zip(header, r))
            if rec.get("Fixture Group") == competition_name:
                rows.append(rec)
    return rows


# ---- Pivot -------------------------------------------------------------------

def pivot_match_rows(raw_rows: list[dict]) -> list[dict]:
    """
    Group raw long-format rows by (player_name, team) within a match.
    Returns one wide dict per (player, team) ready for the import API.
    """
    bucket: dict[tuple, dict] = {}
    unknown_stats: set = set()

    for r in raw_rows:
        player_name = strip_gender(f"{r['First Name']} {r['Last Name']}".strip())
        team_full = strip_team_code(r["Team"])
        key = (player_name.lower(), team_full.lower())

        if key not in bucket:
            bucket[key] = {
                "_player_name": player_name,
                "_team_full": team_full,
            }

        stat = r["Stat Desc"]
        val = r["Stat Value"] or 0
        if stat == "Apps":
            continue
        if stat not in STAT_MAP:
            unknown_stats.add(stat)
            continue
        col = STAT_MAP[stat]
        bucket[key][col] = bucket[key].get(col, 0) + int(val)

    if unknown_stats:
        print(f"  ⚠️  Unknown Stat Desc values ignored: {sorted(unknown_stats)}", file=sys.stderr)
    return list(bucket.values())


# ---- Match lookup ------------------------------------------------------------

def fetch_competition_id(token, base_url, comp_name):
    resp = api_request("GET", f"/admin/competitions?search={urllib.parse.quote(comp_name)}", token, base_url)
    for c in resp.get("data", []):
        if c["name"].strip().lower() == comp_name.strip().lower():
            return c["id"]
    raise SystemExit(f"Competition not found in API: {comp_name!r}. Got: {[c['name'] for c in resp.get('data', [])]}")

def fetch_all_matches(token, base_url, competition_id):
    out = []
    page = 1
    while True:
        resp = api_request("GET", f"/matches?competition_id={competition_id}&page={page}&limit=100", token, base_url)
        data = resp.get("data") or []
        if not data:
            break
        out.extend(data)
        if page >= (resp.get("total_pages") or 1):
            break
        page += 1
    return out

def build_match_index(matches):
    """key: (YYYY-MM-DD, home_name_lower, away_name_lower) -> match_id"""
    idx = {}
    for m in matches:
        home = (m.get("home_team") or {}).get("name", "").strip().lower()
        away = (m.get("away_team") or {}).get("name", "").strip().lower()
        date = (m.get("date") or "")[:10]
        if not (home and away and date):
            continue
        idx[(date, home, away)] = m["id"]
    return idx


# ---- Main --------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="Path to the StatisticDownload xlsx")
    ap.add_argument("--dry-run", action="store_true", help="Parse, pivot, and resolve matches but don't POST")
    ap.add_argument("--only", help="Substring filter on the fixture date string (e.g. 'Sun 24 Sep 2023')")
    ap.add_argument("--competition", default=COMPETITION_NAME, help="Fixture Group value to filter on (default: %(default)r)")
    ap.add_argument("--competition-id", help="DB competition ID. Skips the name-based API lookup; use when the Excel name and DB name differ.")
    ap.add_argument("--sheet", help="Workbook sheet name to read. If omitted, all sheets are scanned for matching rows.")
    args = ap.parse_args()

    token = os.environ.get("SHOWTIME_TOKEN")
    base_url = os.environ.get("SHOWTIME_API_URL", "http://localhost:8089/api/v1").rstrip("/")
    if not token and not args.dry_run:
        raise SystemExit("SHOWTIME_TOKEN env var is required (paste your admin bearer token)")

    raw_rows = read_competition_rows(args.xlsx, args.competition, sheet=args.sheet)
    if args.only:
        raw_rows = [r for r in raw_rows if args.only.lower() in r["Fixture Date"].lower()]
    if not raw_rows:
        raise SystemExit(f"No rows found for {args.competition!r}" + (f" matching {args.only!r}" if args.only else ""))

    # Group raw rows by match: (date_str, home, away)
    by_match = defaultdict(list)
    for r in raw_rows:
        date_iso = parse_fixture_date(r["Fixture Date"]).strftime("%Y-%m-%d")
        home = strip_team_code(r["Home Team"])
        away = strip_team_code(r["Away Team"])
        by_match[(date_iso, home, away)].append(r)

    print(f"Found {len(by_match)} unique matches for {args.competition!r} ({len(raw_rows)} raw rows).")

    # Look up matches in the live DB.
    if not args.dry_run:
        if args.competition_id:
            comp_id = args.competition_id
            print(f"Competition ID (from --competition-id): {comp_id}")
        else:
            comp_id = fetch_competition_id(token, base_url, args.competition)
            print(f"Competition ID: {comp_id}")
        live_matches = fetch_all_matches(token, base_url, comp_id)
        print(f"Live matches in competition: {len(live_matches)}")
        idx = build_match_index(live_matches)
    else:
        idx = {}

    successes, failures, skipped = 0, 0, 0
    for (date_iso, home, away), rows in sorted(by_match.items()):
        # Side resolution stays on the Excel-side names — that's what `p['_team_full']` is.
        # The override is only used when looking up the match in the live DB.
        home_db = TEAM_NAME_OVERRIDES.get(home.lower(), home)
        away_db = TEAM_NAME_OVERRIDES.get(away.lower(), away)
        key = (date_iso, home_db.lower(), away_db.lower())

        pivoted = pivot_match_rows(rows)
        api_rows = []
        for p in pivoted:
            if p["_team_full"].lower() == home.lower():
                side = "home"
            elif p["_team_full"].lower() == away.lower():
                side = "away"
            else:
                print(f"  ⚠️  {date_iso} {home} v {away}: player {p['_player_name']!r} on team {p['_team_full']!r} matches neither side — skipped")
                continue
            api_row = {"side": side, "player_name": p["_player_name"]}
            for v in STAT_MAP.values():
                if v in p:
                    api_row[v] = p[v]
            api_rows.append(api_row)

        if args.dry_run:
            print(f"  [DRY] {date_iso} {home_db} v {away_db}: {len(api_rows)} player-rows")
            successes += 1
            continue

        match_id = idx.get(key)
        if not match_id:
            print(f"  ❌ {date_iso} {home_db} v {away_db}: no matching match in DB — skipped")
            skipped += 1
            continue

        try:
            resp = api_request("POST", f"/admin/matches/{match_id}/import", token, base_url, body={"rows": api_rows})
            d = resp.get("data", {})
            print(
                f"  ✅ {date_iso} {home_db} v {away_db}: "
                f"{d.get('sheet_rows')} sheet, {d.get('stat_rows')} stats, "
                f"{d.get('players_created')} created / {d.get('players_matched')} matched"
            )
            successes += 1
        except SystemExit as e:
            print(f"  ❌ {date_iso} {home_db} v {away_db}: {e}")
            failures += 1

    print(f"\nDone. ✅ {successes}  ❌ {failures}  ⚠️ skipped {skipped}")


if __name__ == "__main__":
    main()
